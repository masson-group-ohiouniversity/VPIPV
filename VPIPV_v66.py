import numpy as np
import math
import streamlit as st
import plotly.graph_objects as go
from scipy.optimize import least_squares
import pandas as pd
import openpyxl
import hashlib
from io import BytesIO

st.set_page_config(page_title="VPIPV Simulator", layout="wide")


# --- Apply fitted parameter updates BEFORE widgets are created (prevents StreamlitAPIException) ---
if st.session_state.get("_pending_fit_param_updates") is not None:
    _upd = st.session_state.pop("_pending_fit_param_updates")
    if isinstance(_upd, dict):
        for _k, _v in _upd.items():
            st.session_state[_k] = _v


from decimal import Decimal, ROUND_FLOOR, ROUND_CEILING

# -----------------------------
# Streamlit widget helpers
# -----------------------------
def _num_input_state(label: str, key: str, default, **kwargs):
    """A number_input that keeps the v44 default on first launch, but avoids
    the Streamlit warning about setting a widget value via Session State while
    also providing a default value.

    Behavior:
      - If the widget key is not yet in st.session_state, initialize it to `default`.
      - Create the widget WITHOUT passing `value=` so Streamlit uses session_state.
    """
    if key not in st.session_state:
        st.session_state[key] = default
    return st.number_input(label, key=key, **kwargs)

# -----------------------------
# Excel import helpers (v45)
# -----------------------------
def _num_input_state_fit(label, key, default, step=0.1, disabled=False):
    # Numeric input + a small "fit" flag beside it (only affects fitting; no effect otherwise)
    c1, c2 = st.columns([0.86, 0.14])
    with c1:
        v = _num_input_state(label, key=key, default=default, step=step, disabled=disabled)
    with c2:
        fit_key = f"fit_{key}"
        if fit_key not in st.session_state:
            st.session_state[fit_key] = False
        st.checkbox("fit", key=fit_key, disabled=disabled)
    return v


def _as_str(x):
    return "" if x is None else str(x).strip()

def _is_no(x):
    s = _as_str(x).lower()
    return s in {"no", "n", "false", "0", ""}

def _is_fit(x):
    s = _as_str(x).lower()
    return s == "fit"

def _to_float_or_none(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = _as_str(x)
    if s == "" or s.lower() == "none":
        return None
    try:
        return float(s)
    except Exception:
        return None

def load_excel_inputs(xlsx_bytes: bytes):
    """Parse the VPIPV Excel input template (Sheet1). Concentrations in mM; V0 and V added in mL."""
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)
    ws = wb["Sheet1"]

    # Top block
    G0_mM = _to_float_or_none(ws["B1"].value)
    H0_mM = _to_float_or_none(ws["B2"].value)

    enable_Q_cell = ws["B3"].value
    enable_Q = not _is_no(enable_Q_cell)

    Q0_mM = _to_float_or_none(ws["B4"].value) if enable_Q else 0.0
    V0_mL = _to_float_or_none(ws["B5"].value)
    Mtitrant_mM = _to_float_or_none(ws["B6"].value)

    # Constants table rows 7–22 (col A name, col B value or 'fit' or 'no')
    const_map = {
        "log KH3M": "logKH3M",
        "log KH3M2": "logKH3M2",
        "log KH1": "logKH1",
        "log KH2": "logKH2",
        "log KH3": "logKH3",
        "log KHM": "logKHM",
        "log KHM2": "logKHM2",
        "log KM,0": "logKM0",
        "log KM2,0": "logKM20",
        "log KM,1": "logKM1",
        "log KM2,1": "logKM21",
        "log KM,2": "logKM2",
        "log KM2,2": "logKM22",
        "log KQ": "logKQ",
        "log KQM": "logKQM",
        "log KQHM": "logKQHM",
    }

    const_vals = {}
    fit_flags = set()
    allow_KH3 = None
    allow_HM = None

    for r in range(7, 23):
        name = _as_str(ws[f"A{r}"].value)
        val = ws[f"B{r}"].value
        if name == "":
            continue
        if name not in const_map:
            continue
        key = const_map[name]

        if name == "log KH3":
            fval = _to_float_or_none(val)
            if _is_no(val) or fval is None:
                allow_KH3 = False
                const_vals[key] = -99.0
            else:
                allow_KH3 = True
                const_vals[key] = float(fval)
            continue

        if name in {"log KHM", "log KHM2"}:
            # Decide allow_HM from presence of either value (unless explicitly "no")
            fval = _to_float_or_none(val)
            if _is_no(val) or fval is None:
                # keep value unset; allow_HM may remain None until we see the other one
                pass
            else:
                allow_HM = True
                const_vals[key] = float(fval)
            continue

        if _is_fit(val):
            fit_flags.add(key)
        else:
            f = _to_float_or_none(val)
            if f is not None:
                const_vals[key] = float(f)

    # finalize allow_HM: yes if any KHM / KHM2 provided; otherwise no
    if allow_HM is None:
        # if either was explicitly provided as a number in const_vals, allow_HM already True
        allow_HM = any(k in const_vals for k in ("logKHM", "logKHM2"))

    # Experimental data (rows 24+): col A = V added (mL), col B = %S3
    v_added = []
    pct_s3 = []
    r = 24
    while True:
        va = ws[f"A{r}"].value
        ys = ws[f"B{r}"].value
        if va is None and ys is None:
            break
        va_f = _to_float_or_none(va)
        ys_f = _to_float_or_none(ys)
        if va_f is not None and ys_f is not None:
            v_added.append(float(va_f))
            pct_s3.append(float(ys_f))
        r += 1
        if r > 5000:
            break

    
    # Enforce toggle rule from template: allow_HM is YES if B12 or B13 is numeric
    allow_HM = (isinstance(ws["B12"].value, (int, float)) or isinstance(ws["B13"].value, (int, float)))
    return {
        "G0_mM": G0_mM,
        "H0_mM": H0_mM,
        "enable_Q": enable_Q,
        "Q0_mM": Q0_mM,
        "V0_mL": V0_mL,
        "Mtitrant_mM": Mtitrant_mM,
        "const_vals": const_vals,
        "fit_flags": sorted(list(fit_flags)),
        "allow_KH3": allow_KH3,
        "allow_HM": bool(allow_HM),
        "v_added_mL": np.array(v_added, dtype=float),
        "pctS3": np.array(pct_s3, dtype=float),
    }

def equiv_from_vadded(v_added_mL, G0_mM, V0_mL, Mtitrant_mM):
    nG0 = float(G0_mM) * float(V0_mL)
    if nG0 <= 0:
        return np.zeros_like(v_added_mL, dtype=float)
    return (float(Mtitrant_mM) * np.array(v_added_mL, dtype=float)) / nG0



def pctS3_from_params_at_equivs(p: dict, equivs: np.ndarray):
    p_local = dict(p)
    p_local["xs_custom"] = np.array(equivs, dtype=float)
    # Force percent mode for fitting target
    p_local["yMode"] = "pct"
    xs, y, S0, S1, S2, S3, Hfree_mM, Mfree_mM, warn, resid_norm = _compute_curve_once(p_local)
    total = S0 + S1 + S2 + S3
    pct_S3 = np.where(total > 0, 100.0 * S3 / total, 0.0)
    return xs, pct_S3

def fit_logKs_to_pctS3(p: dict, fit_keys: list, equivs: np.ndarray, pctS3_exp: np.ndarray):
    """Fit selected logK parameters (given by keys in p) to experimental %S3 vs equiv."""
    fit_keys = list(fit_keys)
    x0 = np.array([float(p[k]) for k in fit_keys], dtype=float)

    def unpack(x):
        p2 = dict(p)
        for k, v in zip(fit_keys, x):
            p2[k] = float(v)
        return p2

    def residuals(x):
        p2 = unpack(x)
        _, pct_pred = pctS3_from_params_at_equivs(p2, equivs)
        return (pct_pred - pctS3_exp)

    # broad but safe bounds in log10 space
    lo = np.full_like(x0, -10.0, dtype=float)
    hi = np.full_like(x0,  20.0, dtype=float)

    res = least_squares(residuals, x0, bounds=(lo, hi), method="trf")
    p_fit = unpack(res.x)
    return p_fit, res

def snap_logk_input(label: str, key: str, value: float, disabled: bool = False):
    """LogK input with minimal custom +/- that snaps to the nearest 0.1 grid before stepping.
    Only used for equilibrium constants to avoid 'buttons all over'.
    """
    if key not in st.session_state:
        st.session_state[key] = float(value)

    def _to_dec(x: float) -> Decimal:
        return Decimal(str(x))

    def _on_grid(x: float) -> bool:
        d = _to_dec(x)
        return (d * Decimal("10")) == (d * Decimal("10")).to_integral_value()

    def _snap_up(x: float) -> float:
        d = _to_dec(x)
        return float(((d * Decimal("10")).to_integral_value(rounding=ROUND_CEILING)) / Decimal("10"))

    def _snap_down(x: float) -> float:
        d = _to_dec(x)
        return float(((d * Decimal("10")).to_integral_value(rounding=ROUND_FLOOR)) / Decimal("10"))

    c1, c2, c3 = st.columns([8, 1, 1], gap="small")
    with c1:
        v = st.number_input(label, value=float(st.session_state[key]), step=0.0, format="%.2f",
                            disabled=disabled, key=f"{key}__box")
        st.session_state[key] = float(v)

    with c2:
        if st.button("−", key=f"{key}__minus", disabled=disabled):
            x = float(st.session_state[key])
            x_new = (x - 0.1) if _on_grid(x) else _snap_down(x)
            st.session_state[key] = float(x_new)

    with c3:
        if st.button("+", key=f"{key}__plus", disabled=disabled):
            x = float(st.session_state[key])
            x_new = (x + 0.1) if _on_grid(x) else _snap_up(x)
            st.session_state[key] = float(x_new)

    return float(st.session_state[key])


# -----------------------------
# Chemistry model helpers
# -----------------------------
def pow10(x: float) -> float:
    return 10.0 ** x

def coeffs_G(H: float, M: float, p: dict) -> dict:
    """
    Coefficients multiplying G_free for each G-species:
      [species] = coeff * G_free

    Network (11 equilibria + optional GH3):
      G, GM, GM2
      GH, GHM, GHM2
      GH2, GH2M, GH2M2
      GH3 (optional), GH3M, GH3M2

    with special steps:
      GH2M  + H <-> GH3M   (KH3M)
      GH2M2 + H <-> GH3M2  (KH3M2)
    """
    KM0  = pow10(p["logKM0"])
    KM20 = pow10(p["logKM20"])
    KM1  = pow10(p["logKM1"])
    KM21 = pow10(p["logKM21"])
    KM2  = pow10(p["logKM2"])
    KM22 = pow10(p["logKM22"])

    KH1  = pow10(p["logKH1"])
    KH2  = pow10(p["logKH2"])
    KH3  = pow10(p["logKH3"]) if p["allow_KH3"] else 0.0

    KH3M  = pow10(p["logKH3M"])
    KH3M2 = pow10(p["logKH3M2"])

    # Level 0
    c_G   = 1.0
    c_GM  = KM0 * M
    c_GM2 = KM0 * KM20 * M * M

    # Level 1
    c_GH   = KH1 * H
    c_GHM  = c_GH * KM1 * M
    c_GHM2 = c_GH * KM1 * KM21 * M * M

    # Level 2
    c_GH2   = (KH1 * KH2) * H * H
    c_GH2M  = c_GH2 * KM2 * M
    c_GH2M2 = c_GH2 * KM2 * KM22 * M * M

    # Level 3 (optional metal-free GH3)
    c_GH3 = 0.0 if KH3 == 0.0 else (KH1 * KH2 * KH3) * H * H * H

    # Metalated level-3 via special steps
    c_GH3M  = c_GH2M  * KH3M  * H
    c_GH3M2 = c_GH2M2 * KH3M2 * H

    return {
        "G": c_G,
        "GM": c_GM,
        "GM2": c_GM2,
        "GH": c_GH,
        "GHM": c_GHM,
        "GHM2": c_GHM2,
        "GH2": c_GH2,
        "GH2M": c_GH2M,
        "GH2M2": c_GH2M2,
        "GH3": c_GH3,
        "GH3M": c_GH3M,
        "GH3M2": c_GH3M2,
    }

def q_partition(H: float, M: float, Qtot: float, p: dict):
    """
    Optional competitor Q network:
      Q + H  <-> QH   with KQ
      Q + M  <-> QM   with KQM
      QH + M <-> QHM  with KQHM

    Binding polynomial: denom = 1 + KQ*H + KQM*M + (KQ*KQHM)*H*M
    """
    if (not p["enable_Q"]) or Qtot <= 0.0:
        return dict(Qfree=Qtot, QH=0.0, QM=0.0, QHM=0.0, HboundQ=0.0, MboundQ=0.0)

    KQ   = pow10(p["logKQ"])
    KQM  = pow10(p["logKQM"])
    KQHM = pow10(p["logKQHM"])

    denom = 1.0 + KQ * H + KQM * M + (KQ * KQHM) * H * M
    Qfree = Qtot / denom
    QH  = KQ * Qfree * H
    QM  = KQM * Qfree * M
    QHM = (KQ * KQHM) * Qfree * H * M

    HboundQ = QH + QHM
    MboundQ = QM + QHM
    return dict(Qfree=Qfree, QH=QH, QM=QM, QHM=QHM, HboundQ=HboundQ, MboundQ=MboundQ)

def solve_free_HM(Gtot: float, Htot: float, Mtot: float, Qtot: float, p: dict, x0_lnHM: np.ndarray):
    """
    Solve mass balances for free H and free M using SciPy least_squares in log-space.
    Variables: lnH, lnM (so H,M are positive).
    """
    def residuals(x):
        lnH, lnM = x
        H = np.exp(lnH)
        M = np.exp(lnM)

        c = coeffs_G(H, M, p)
        Gfree = Gtot / sum(c.values())

        HboundG = Gfree * (
            1*(c["GH"] + c["GHM"] + c["GHM2"]) +
            2*(c["GH2"] + c["GH2M"] + c["GH2M2"]) +
            3*(c["GH3"] + c["GH3M"] + c["GH3M2"])
        )

        MboundG = Gfree * (
            1*(c["GM"] + c["GHM"] + c["GH2M"] + c["GH3M"]) +
            2*(c["GM2"] + c["GHM2"] + c["GH2M2"] + c["GH3M2"])
        )

        q = q_partition(H, M, Qtot, p)
        HboundQ = q["HboundQ"]
        MboundQ = q["MboundQ"]

        # Optional direct host–metal binding (does not contribute to S0–S3 bins)
        KHM  = pow10(p["logKHM"]) if p.get("allow_HM", False) else 0.0
        KHM2 = pow10(p["logKHM2"]) if p.get("allow_HM", False) else 0.0
        HM  = KHM * H * M
        HM2 = (KHM * KHM2) * H * M * M
        HboundHM = HM + HM2
        MboundHM = HM + 2.0 * HM2

        return np.array([
            (H + HboundG + HboundQ + HboundHM) - Htot,
            (M + MboundG + MboundQ + MboundHM) - Mtot
        ])

    lo = np.array([-46.0, -46.0])  # ~1e-20 M
    hi = np.array([  0.0,   0.0])  # 1 M

    sol = least_squares(
        residuals,
        x0=x0_lnHM,
        bounds=(lo, hi),
        xtol=1e-14, ftol=1e-14, gtol=1e-14,
        max_nfev=6000
    )

    lnH, lnM = sol.x
    H = float(np.exp(lnH))
    M = float(np.exp(lnM))
    return sol, H, M


def compute_curve(p: dict):
    """Compute curve; if warning points exist, increase nPts by 1.3x and retry until warnings clear (capped)."""
    nPts_start = int(p["nPts"])
    max_iter = int(p.get("autoRefineMaxIter", 8))
    max_pts = int(p.get("autoRefineMaxPts", 2000))
    grow = float(p.get("autoRefineGrow", 1.3))

    nPts_eff = max(10, nPts_start)
    last = None

    for k in range(max_iter):
        p_local = dict(p)
        p_local["nPts"] = int(nPts_eff)
        last = _compute_curve_once(p_local)
        xs, y, S0, S1, S2, S3, Hfree_mM, Mfree_mM, warn, resid_norm = last
        if (warn is not None) and (not bool(np.any(warn))):
            break
        # grow and retry
        nPts_next = int(math.ceil(nPts_eff * grow))
        if nPts_next <= nPts_eff:
            nPts_next = nPts_eff + 1
        nPts_eff = nPts_next
        if nPts_eff > max_pts:
            break
    # diagnostics
    st.session_state["nPts_used"] = int(nPts_eff) if last is not None else nPts_start
    return last


def _compute_curve_once(p: dict):
    """
    Build titration curve vs equiv Ag0 with dilution.
    Equivalent definition:
      equiv = (moles Ag added) / (initial moles G)
    """
    # initial amounts in mmol
    nG0 = p["G0_mM"] * p["V0_mL"]
    nH0 = p["H0_mM"] * p["V0_mL"]
    nQ0 = p["Q0_mM"] * p["V0_mL"]

    Mtitrant = max(1e-12, p["Mtitrant_mM"])
    xs = np.array(p.get("xs_custom"), dtype=float) if p.get("xs_custom") is not None else np.linspace(0.0, float(p["maxEquiv"]), int(p["nPts"]))

    y = np.zeros_like(xs)
    S0 = np.zeros_like(xs)
    S1 = np.zeros_like(xs)
    S2 = np.zeros_like(xs)
    S3 = np.zeros_like(xs)
    Hfree_mM = np.zeros_like(xs)
    Mfree_mM = np.zeros_like(xs)
    warn = np.zeros_like(xs, dtype=bool)
    resid_norm = np.zeros_like(xs)

    # heuristic initial guess
    H_guess = max((p["H0_mM"] - 2*p["G0_mM"] - (p["Q0_mM"] if p["enable_Q"] else 0.0)) * 1e-3, 1e-15)
    M_guess = 1e-15
    x0 = np.log([H_guess, M_guess])

    for i, eq in enumerate(xs):
        nMadd = eq * nG0   # mmol
        Vadd = nMadd / Mtitrant  # mL
        V = p["V0_mL"] + Vadd

        Gtot = (nG0 / V) * 1e-3  # M
        Htot = (nH0 / V) * 1e-3  # M
        Qtot = (nQ0 / V) * 1e-3  # M
        Mtot = (nMadd / V) * 1e-3  # M

        if i == 0:
            x0 = np.log([max(Htot, 1e-15), max(Mtot, 1e-15)])

        sol, H, M = solve_free_HM(Gtot, Htot, Mtot, Qtot, p, x0)
        x0 = sol.x  # continuation
        resid_norm[i] = float(np.linalg.norm(sol.fun))

        warn[i] = (not sol.success) or (resid_norm[i] > 1e-10)

        c = coeffs_G(H, M, p)
        Gfree = Gtot / sum(c.values())

        S0_i = c["G"] * Gfree
        S1_i = (c["GH"] + c["GHM"] + c["GHM2"]) * Gfree
        S2_i = (c["GH2"] + c["GH2M"] + c["GH2M2"]) * Gfree
        S3_i = (c["GH3"] + c["GH3M"] + c["GH3M2"]) * Gfree

        S0[i] = S0_i * 1e3
        S1[i] = S1_i * 1e3
        S2[i] = S2_i * 1e3
        S3[i] = S3_i * 1e3
        Hfree_mM[i] = H * 1e3
        Mfree_mM[i] = M * 1e3

        if p["yMode"] == "pct":
            y[i] = 100.0 * (S3_i / Gtot)
        elif p["yMode"] == "S3mM":
            y[i] = S3[i]
        elif p["yMode"] == "freeH":
            y[i] = Hfree_mM[i]
        elif p["yMode"] == "freeM":
            y[i] = Mfree_mM[i]
        else:
            y[i] = 100.0 * (S3_i / Gtot)

    return xs, y, S0, S1, S2, S3, Hfree_mM, Mfree_mM, warn, resid_norm

# -----------------------------
# UI
# -----------------------------
st.title("VPIPV Simulator")

with st.sidebar:
    uploaded_xlsx = st.file_uploader("Open File", type=["xlsx"])
    # Persist uploaded Excel across reruns (needed for fitting reruns)
    if uploaded_xlsx is not None:
        st.session_state["_excel_xlsx_bytes"] = uploaded_xlsx.getvalue()
    elif st.session_state.get("_excel_xlsx_bytes") is not None:
        class _MemUpload:
            def __init__(self, b):
                self._b = b
            def getvalue(self):
                return self._b
        uploaded_xlsx = _MemUpload(st.session_state["_excel_xlsx_bytes"])

    if uploaded_xlsx is not None:
        try:
            xlsx_bytes = uploaded_xlsx.getvalue()
            excel_inputs = load_excel_inputs(xlsx_bytes)
            excel_digest = hashlib.md5(xlsx_bytes).hexdigest()
            st.session_state["excel_loaded"] = True
            st.session_state["excel_v_added_mL"] = excel_inputs["v_added_mL"]
            st.session_state["excel_pctS3"] = excel_inputs["pctS3"]

            if len(excel_inputs["fit_flags"]) > 0:
                if st.button("Fit constants marked 'fit'"):
                    st.session_state["do_fit"] = True
            if st.session_state.get("excel_digest") != excel_digest:
                # Clear any prior widget choices so Excel can drive them
                st.session_state.pop("enable_Q_choice", None)
                st.session_state.pop("allow_HM_choice", None)
                # Fit-flag initialization (only on new Excel upload)
                st.session_state["excel_fit_flags"] = excel_inputs["fit_flags"]
                for _k in ["logKH3M","logKH3M2","logKH1","logKH2","logKH3","logKHM","logKHM2","logKM0","logKM20","logKM1","logKM21","logKM2","logKM22","logKQ","logKQM","logKQHM"]:
                    st.session_state[f"fit_{_k}"] = (_k in st.session_state.get("excel_fit_flags", []))

                # Core concentrations / toggles
                if excel_inputs["G0_mM"] is not None:
                    st.session_state["G0_mM"] = float(excel_inputs["G0_mM"])
                if excel_inputs["H0_mM"] is not None:
                    st.session_state["H0_mM"] = float(excel_inputs["H0_mM"])
                if excel_inputs["V0_mL"] is not None:
                    st.session_state["V0_mL"] = float(excel_inputs["V0_mL"])
                if excel_inputs["Mtitrant_mM"] is not None:
                    st.session_state["Mtitrant_mM"] = float(excel_inputs["Mtitrant_mM"])

                st.session_state["enable_Q"] = bool(excel_inputs["enable_Q"])
                st.session_state["enable_Q_choice"] = "Yes" if st.session_state["enable_Q"] else "No"
                if excel_inputs["enable_Q"]:
                    # Only overwrite Q0 if Excel provided a numeric value
                    if excel_inputs["Q0_mM"] is not None:
                        st.session_state["Q0_mM"] = float(excel_inputs["Q0_mM"])
                else:
                    st.session_state["Q0_mM"] = 0.0

                st.session_state["allow_KH3"] = bool(excel_inputs["allow_KH3"]) if excel_inputs["allow_KH3"] is not None else False
                st.session_state["allow_HM"] = bool(excel_inputs["allow_HM"])
                st.session_state["allow_HM_choice"] = "Yes" if st.session_state["allow_HM"] else "No (KHM=0)"

                # Constants
                for k, v in excel_inputs["const_vals"].items():
                    st.session_state[k] = float(v)

                st.session_state["excel_digest"] = excel_digest
                st.rerun()
        except Exception as e:
            st.error(f"Could not parse Excel file: {e}")
    st.header("Initial concentrations (mM)")
    G0_mM = _num_input_state("G0 (guest)", key="G0_mM", default=1.0, step=0.1, format="%.3f")
    H0_mM = _num_input_state("H0 (host)", key="H0_mM", default=3.1, step=0.1, format="%.3f")
    _enable_Q_label = st.selectbox("Enable competitor Q?", ["Yes", "No"], key="enable_Q_choice")
    enable_Q = (_enable_Q_label == "Yes")
    st.session_state["enable_Q"] = enable_Q
    Q0_mM = _num_input_state("Q0 (competitor)", key="Q0_mM", default=1.0, step=0.1, format="%.3f", disabled=not enable_Q)

    st.header("Titration / plotting")
    V0_mL = _num_input_state("V0 (mL)", key="V0_mL", default=0.5, step=0.01, format="%.3f")
    Mtitrant_mM = _num_input_state("Ag titrant (mM)", key="Mtitrant_mM", default=10.0, step=0.1, format="%.3f")
    xMaxBox = _num_input_state("X-axis max (0 to X)", key="xMaxBox", default=3.0, step=0.1, format="%.3f")
    maxEquiv = float(xMaxBox)  # <-- hard override
    nPts = st.number_input("# points", value=61, step=1)

    yMode_label = st.selectbox("Y axis", ["%S", "S [mM]"], index=0)
    showBins = st.selectbox("Show bins S0–S3?", ["No", "Yes"], index=0) == "Yes"

    st.header("Binding constants (log10 K)")
    logKH3M  = _num_input_state_fit("log KH3M (GH2M + H ⇌ GH3M)", key="logKH3M", default=5.0, step=0.1)
    logKH3M2 = _num_input_state_fit("log KH3M2 (GH2M2 + H ⇌ GH3M2)", key="logKH3M2", default=5.0, step=0.1)
    logKH1   = _num_input_state_fit("log KH1 (G + H ⇌ GH)", key="logKH1", default=5.86, step=0.1)
    logKH2   = _num_input_state_fit("log KH2 (GH + H ⇌ GH2)", key="logKH2", default=5.86, step=0.1)

    _allow_KH3_label = st.selectbox("Allow GH3 (KH3)?", ["No (KH3=0)", "Yes"], index=(1 if bool(st.session_state.get("allow_KH3", False)) else 0), key="allow_KH3_choice")
    allow_KH3 = (_allow_KH3_label == "Yes")
    st.session_state["allow_KH3"] = allow_KH3
    logKH3 = _num_input_state_fit("log KH3 (GH2 + H ⇌ GH3)", key="logKH3", default=-99.0, step=0.1, disabled=not allow_KH3)

    _allow_HM_label = st.selectbox("H binds to M?", ["No (KHM=0)", "Yes"], key="allow_HM_choice")
    allow_HM = (_allow_HM_label == "Yes")
    st.session_state["allow_HM"] = allow_HM
    logKHM  = _num_input_state_fit("log KHM (H + M ⇌ HM)", key="logKHM", default=3.54, step=0.1, disabled=not allow_HM)
    logKHM2 = _num_input_state_fit("log KHM2 (HM + M ⇌ HM2)", key="logKHM2", default=3.54, step=0.1, disabled=not allow_HM)

    logKM0  = _num_input_state_fit("log KM,0 (G + M ⇌ GM)", key="logKM0", default=4.23, step=0.1)
    logKM20  = _num_input_state_fit("log KM2,0 (GM + M ⇌ GM2)", key="logKM20", default=4.23, step=0.1)
    logKM1  = _num_input_state_fit("log KM,1 (GH + M ⇌ GHM)", key="logKM1", default=4.23, step=0.1)
    logKM21  = _num_input_state_fit("log KM2,1 (GHM + M ⇌ GHM2)", key="logKM21", default=4.23, step=0.1)
    logKM2  = _num_input_state_fit("log KM,2 (GH2 + M ⇌ GH2M)", key="logKM2", default=4.23, step=0.1)
    logKM22  = _num_input_state_fit("log KM2,2 (GH2M + M ⇌ GH2M2)", key="logKM22", default=4.23, step=0.1)

    st.header("Competitor Q (log10 K)")
    logKQ   = _num_input_state_fit("log KQ (Q + H ⇌ QH)", key="logKQ", default=4.0, step=0.1, disabled=not enable_Q)
    logKQM  = _num_input_state_fit("log KQM (Q + M ⇌ QM)", key="logKQM", default=3.0, step=0.1, disabled=not enable_Q)
    logKQHM = _num_input_state_fit("log KQHM (QH + M ⇌ QHM)", key="logKQHM", default=5.0, step=0.1, disabled=not enable_Q)

yMode_map = {"%S":"pct", "S [mM]":"S3mM"}

params = dict(
    G0_mM=float(G0_mM),
    H0_mM=float(H0_mM),
    Q0_mM=float(Q0_mM) if enable_Q else 0.0,
    enable_Q=bool(enable_Q),
    V0_mL=float(V0_mL),
    Mtitrant_mM=float(Mtitrant_mM),
    maxEquiv=float(maxEquiv),
    nPts=int(nPts),
    yMode=yMode_map[yMode_label],
    showBins=bool(showBins),
    logKH3M=float(logKH3M),
    logKH3M2=float(logKH3M2),
    logKH1=float(logKH1),
    logKH2=float(logKH2),
    allow_KH3=bool(allow_KH3),
    logKH3=float(logKH3),
    allow_HM=bool(allow_HM),
    logKHM=float(logKHM),
    logKHM2=float(logKHM2),
    logKM0=float(logKM0),
    logKM20=float(logKM20),
    logKM1=float(logKM1),
    logKM21=float(logKM21),
    logKM2=float(logKM2),
    logKM22=float(logKM22),
    logKQ=float(logKQ),
    logKQM=float(logKQM),
    logKQHM=float(logKQHM),
)


# --- Optional experimental data from Excel ---
exp_equiv = None
exp_pctS3 = None
exp_S3_mM = None
if st.session_state.get("excel_loaded", False):
    v_added = st.session_state.get("excel_v_added_mL", None)
    pct = st.session_state.get("excel_pctS3", None)
    if v_added is not None and pct is not None and len(v_added) > 0:
        exp_equiv = equiv_from_vadded(v_added, params["G0_mM"], params["V0_mL"], params["Mtitrant_mM"])
        exp_pctS3 = np.array(pct, dtype=float)
        exp_S3_mM = (exp_pctS3/100.0) * (float(params["G0_mM"])*float(params["V0_mL"])) / (float(params["V0_mL"]) + np.array(v_added, dtype=float))

# --- Optional fitting (Excel constants marked 'fit') ---
if st.session_state.get("do_fit", False):
    fit_keys = [k for k in ["logKH3M","logKH3M2","logKH1","logKH2","logKH3","logKHM","logKHM2","logKM0","logKM20","logKM1","logKM21","logKM2","logKM22","logKQ","logKQM","logKQHM"] if st.session_state.get(f"fit_{k}", False)]
    if exp_equiv is None or exp_pctS3 is None or len(fit_keys) == 0:
        st.sidebar.error("Fit requested, but missing experimental data or no constants marked 'fit' in Excel.")
        st.session_state["do_fit"] = False
    else:
        p_fit, res = fit_logKs_to_pctS3(params, fit_keys, exp_equiv, exp_pctS3)
        # Defer widget updates to next rerun (cannot set widget keys after instantiation)
        st.session_state["_pending_fit_param_updates"] = {k: float(p_fit[k]) for k in fit_keys}
        st.session_state["fit_report"] = {
            "fit_keys": list(fit_keys),
            "x0": [float(params[k]) for k in fit_keys],
            "x_opt": [float(p_fit[k]) for k in fit_keys],
            "cost": float(res.cost),
            "rmse_pctS3": float(np.sqrt(np.mean(res.fun**2))) if res.fun.size else float("nan"),
            "success": bool(res.success),
            "message": str(res.message),
            "nfev": int(res.nfev),
        }
        st.session_state["do_fit"] = False
        st.rerun()

xs, y, S0, S1, S2, S3, Hfree_mM, Mfree_mM, warn, resid_norm = compute_curve(params)

# --- PLOTTING ---
col1, col2 = st.columns([2.2, 1.0], gap="large")

with col1:
    fig = go.Figure()

    # Plot S0–S3 as percentages
    total = S0 + S1 + S2 + S3
    pct_S0 = np.where(total > 0, 100*S0/total, 0)
    pct_S1 = np.where(total > 0, 100*S1/total, 0)
    pct_S2 = np.where(total > 0, 100*S2/total, 0)
    pct_S3 = np.where(total > 0, 100*S3/total, 0)

    if params["yMode"] == "S3mM":
        # S [mM] mode: plot diluted concentrations (already in mM arrays)
        fig.add_trace(go.Scatter(x=xs, y=S0, mode="lines",
                                 name="S0 (mM)", line=dict(dash="dot")))
        fig.add_trace(go.Scatter(x=xs, y=S1, mode="lines",
                                 name="S1 (mM)", line=dict(dash="dash")))
        fig.add_trace(go.Scatter(x=xs, y=S2, mode="lines",
                                 name="S2 (mM)", line=dict(dash="longdash")))
        fig.add_trace(go.Scatter(x=xs, y=S3, mode="lines",
                                 name="S3 (mM)", line=dict(width=3)))
    else:
        # %S mode
        fig.add_trace(go.Scatter(x=xs, y=pct_S0, mode="lines",
                                 name="S0 (%)", line=dict(dash="dot")))
        fig.add_trace(go.Scatter(x=xs, y=pct_S1, mode="lines",
                                 name="S1 (%)", line=dict(dash="dash")))
        fig.add_trace(go.Scatter(x=xs, y=pct_S2, mode="lines",
                                 name="S2 (%)", line=dict(dash="longdash")))
        fig.add_trace(go.Scatter(x=xs, y=pct_S3, mode="lines",
                                 name="S3 (%)", line=dict(width=3)))

    if np.any(warn):
        fig.add_trace(go.Scatter(
            x=xs[warn], y=(S3[warn] if params["yMode"]=="S3mM" else pct_S3[warn]),
            mode="markers", marker=dict(symbol="x", size=10),
            name="solver warning points"
        ))
    # Experimental points from Excel (if provided): column B is %S3
    if exp_equiv is not None and exp_pctS3 is not None:
        y_exp = exp_S3_mM if params["yMode"]=="S3mM" else exp_pctS3
        fig.add_trace(go.Scatter(
            x=exp_equiv, y=y_exp,
            mode="markers",
            name="Excel data",
            marker=dict(size=9, symbol="circle-open")
        ))

    fig.update_layout(
        height=740,
        margin=dict(l=40, r=20, t=40, b=40),
        xaxis=dict(title="equiv Ag0", range=[0, float(maxEquiv)], autorange=False),
        yaxis=(dict(title="Population (%)", range=[0, 105], tickmode="array", tickvals=list(range(0, 101, 10))) if params["yMode"]=="pct" else dict(title="S [mM]", range=[0, float(params["G0_mM"]) ])),
        template="plotly_dark",
        uirevision=f"xmax={float(maxEquiv):.6f}",
        legend=dict(x=0.01, y=0.99),
    )

    st.plotly_chart(fig, use_container_width=True, key=f"chart_xmax_{float(maxEquiv):.6f}")

with col2:
    st.subheader("Quick readouts")
    st.write(f"**Solver warning points:** {int(np.sum(warn))} / {len(xs)}")

    fit_report = st.session_state.get("fit_report", None)
    if fit_report is not None:
        st.subheader("Fit report")
        st.write(f"**Success:** {fit_report.get('success', False)}")
        st.write(f"**RMSE (%S3):** {fit_report.get('rmse_pctS3', float('nan')):.4g}")
        st.write(f"**nfev:** {fit_report.get('nfev', 0)}")
        for k, x0, x1 in zip(fit_report.get("fit_keys", []), fit_report.get("x0", []), fit_report.get("x_opt", [])):
            st.write(f"{k}: {x0:.4g} → **{x1:.4g}**")

    imax = int(np.argmax(pct_S3))
    st.write(f"**Peak S3:** {pct_S3[imax]:.2f}% at equiv **{xs[imax]:.3g}**")
    st.write(f"**Final S3:** {pct_S3[-1]:.2f}% at equiv **{xs[-1]:.3g}**")

    st.subheader("Free species (last point)")
    st.write(f"free [H] ≈ **{Hfree_mM[-1]:.4g} mM**")
    st.write(f"free [M] ≈ **{Mfree_mM[-1]:.4g} mM**")

    st.subheader("Bins at last point (mM)")
    st.write(f"S0 = {S0[-1]:.4g}")
    st.write(f"S1 = {S1[-1]:.4g}")
    st.write(f"S2 = {S2[-1]:.4g}")
    st.write(f"S3 = {S3[-1]:.4g}")
